import os
import shutil
import zipfile
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime
from dataclasses import dataclass
from docx import Document
from docx.shared import Pt, RGBColor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from core.project_info import ProjectInfo
from core.material_list import MaterialList, MaterialItem
from core.content_validator import ValidationResult
from config import OUTPUT_DIR, PLATFORM_NAMING_RULE


@dataclass
class SubmissionPreview:
    files_to_include: List[Dict]
    missing_required: List[MaterialItem]
    missing_optional: List[MaterialItem]
    sensitive_warnings: List[Dict]
    validation_errors: List[str]
    validation_warnings: List[str]
    can_submit: bool

    def to_dict(self) -> dict:
        return {
            "files_to_include": self.files_to_include,
            "missing_required": [m.to_dict() for m in self.missing_required],
            "missing_optional": [m.to_dict() for m in self.missing_optional],
            "sensitive_warnings": self.sensitive_warnings,
            "validation_errors": self.validation_errors,
            "validation_warnings": self.validation_warnings,
            "can_submit": self.can_submit
        }


class PackageOutput:
    def __init__(self, project_info: ProjectInfo, material_list: MaterialList,
                 generated_files: Dict, validation_result: ValidationResult = None):
        self.project = project_info
        self.materials = material_list
        self.generated_files = generated_files
        self.validation_result = validation_result
        self.output_dir = OUTPUT_DIR / project_info.project_id
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def get_submission_preview(self) -> SubmissionPreview:
        files_to_include = []
        file_mapping = self._collect_confirmed_files()

        current_batch = self.materials.get_current_batch()
        confirmed_codes = set()
        if current_batch and current_batch.confirmed_files:
            for cf in current_batch.confirmed_files:
                for item in self.materials.items:
                    if item.file_path == cf or (item.code in self.generated_files and self.generated_files[item.code] == cf):
                        confirmed_codes.add(item.code)
                        break

        for file_info in file_mapping:
            try:
                original_path = Path(file_info["original_path"])
                file_size = original_path.stat().st_size if original_path.exists() else 0
                size_str = self._format_file_size(file_size)
                source = "模板生成" if file_info.get("is_generated", False) else "用户上传"
                confirmed = file_info["material_code"] in confirmed_codes

                files_to_include.append({
                    "material_code": file_info["material_code"],
                    "material_name": file_info["material_name"],
                    "original_path": file_info["original_path"],
                    "original_name": original_path.name,
                    "name": original_path.name,
                    "is_generated": file_info.get("is_generated", False),
                    "source": source,
                    "file_size": file_size,
                    "size": size_str,
                    "confirmed": confirmed,
                    "exists": original_path.exists()
                })
            except Exception as e:
                print(f"处理文件预览时出错: {file_info.get('original_path', 'unknown')}, {e}")
                continue

        missing_required = []
        missing_optional = []
        try:
            missing_required = self.materials.get_missing_required()
            all_missing = self.materials.get_all_missing()
            missing_optional = [m for m in all_missing if not m.required]
        except Exception as e:
            print(f"获取缺失材料时出错: {e}")

        sensitive_warnings = []
        validation_errors = []
        validation_warnings = []

        if self.validation_result:
            try:
                sensitive_hits = getattr(self.validation_result, 'sensitive_hits', [])
                sensitive_warnings = []
                for hit in sensitive_hits:
                    sensitive_warnings.append({
                        "file": hit.get("file", hit.get("location", "unknown")),
                        "words": hit.get("words", [hit.get("word", "")]),
                        "line": hit.get("line", 0),
                        "context": hit.get("context", "")
                    })

                validation_errors = list(getattr(self.validation_result, 'errors', []))
                all_warnings = list(getattr(self.validation_result, 'warnings', []))
                validation_warnings = [
                    w for w in all_warnings
                    if not any(h.get("word", "") in w for h in sensitive_hits)
                ]
            except Exception as e:
                print(f"处理校验结果时出错: {e}")

        can_submit = False
        try:
            can_submit = (
                len(missing_required) == 0
                and len(files_to_include) > 0
                and all(f.get("exists", False) for f in files_to_include)
            )
        except Exception as e:
            print(f"计算can_submit时出错: {e}")

        return SubmissionPreview(
            files_to_include=files_to_include,
            missing_required=missing_required,
            missing_optional=missing_optional,
            sensitive_warnings=sensitive_warnings,
            validation_errors=validation_errors,
            validation_warnings=validation_warnings,
            can_submit=can_submit
        )

    def _format_file_size(self, size_bytes: int) -> str:
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        else:
            return f"{size_bytes / (1024 * 1024):.1f} MB"

    def build_package(self, ignore_errors: bool = False, use_batch: bool = True) -> Dict:
        result = {
            "success": False,
            "output_files": [],
            "zip_path": "",
            "checklist_path": "",
            "errors": [],
            "preview": None
        }

        preview = self.get_submission_preview()
        result["preview"] = preview.to_dict()

        if not ignore_errors and not preview.can_submit:
            if preview.missing_required:
                result["errors"].append(
                    f"缺少必填材料: {', '.join([m.name for m in preview.missing_required])}"
                )
            if not preview.files_to_include:
                result["errors"].append("没有可打包的文件")
            if result["errors"]:
                return result

        file_mapping = self._collect_confirmed_files() if use_batch else self._collect_all_files()
        if not file_mapping:
            result["errors"].append("没有可打包的文件")
            return result

        renamed_files = self._rename_files_by_platform_rule(file_mapping)
        checklist_path = self._generate_submission_checklist(renamed_files, file_mapping)
        result["checklist_path"] = str(checklist_path)
        renamed_files.append(str(checklist_path))
        zip_path = self._create_zip_package(renamed_files)
        result["zip_path"] = str(zip_path)
        result["output_files"] = [str(f) for f in renamed_files]
        result["file_count"] = len(renamed_files)
        result["success"] = True
        return result

    def _collect_confirmed_files(self) -> List[Dict]:
        collected = []
        current_batch = self.materials.get_current_batch()

        if current_batch and current_batch.confirmed_files:
            for confirmed_file in current_batch.confirmed_files:
                path_obj = Path(confirmed_file)
                if not path_obj.exists():
                    continue

                mat_code = None
                mat_item = None
                is_generated = False

                for key, filepath in self.generated_files.items():
                    if filepath == confirmed_file:
                        mat_code = self._get_material_code_by_key(key)
                        mat_item = self.materials.get_item(mat_code)
                        is_generated = True
                        break

                if not mat_code:
                    for item in self.materials.get_provided():
                        if item.file_path == confirmed_file:
                            mat_code = item.code
                            mat_item = item
                            is_generated = item.generated
                            break

                if not mat_code:
                    for code in ["CPMS", "SQMS", "YLDQ"]:
                        if code in path_obj.name:
                            mat_code = code
                            mat_item = self.materials.get_item(code)
                            is_generated = mat_item.generated if mat_item else True
                            break

                if mat_code and mat_item:
                    collected.append({
                        "original_path": confirmed_file,
                        "material_code": mat_code,
                        "material_name": mat_item.name,
                        "is_generated": is_generated
                    })
            return collected

        return self._collect_all_files()

    def _collect_all_files(self) -> List[Dict]:
        collected = []
        for key, filepath in self.generated_files.items():
            if key.endswith("_error") or not filepath:
                continue
            mat_code = self._get_material_code_by_key(key)
            mat_item = self.materials.get_item(mat_code)
            collected.append({
                "original_path": filepath,
                "material_code": mat_code,
                "material_name": mat_item.name if mat_item else key,
                "is_generated": True
            })
        for item in self.materials.get_provided():
            if item.generated or not item.file_path:
                continue
            if Path(item.file_path).exists():
                collected.append({
                    "original_path": item.file_path,
                    "material_code": item.code,
                    "material_name": item.name,
                    "is_generated": False
                })
        return collected

    def _get_material_code_by_key(self, key: str) -> str:
        mapping = {
            "product_desc": "CPMS",
            "auth_desc": "SQMS",
            "sample_fields": "YLDQ"
        }
        return mapping.get(key, "QTLY")

    def _rename_files_by_platform_rule(self, files: List[Dict]) -> List[str]:
        renamed = []
        date_str = datetime.now().strftime("%Y%m%d")
        scene = self._safe_filename(self.project.trading_scene[:4] if self.project.trading_scene else "场景")
        code = self._safe_filename(self.project.product_code if self.project.product_code else self.project.project_id)
        for idx, file_info in enumerate(files):
            original_path = Path(file_info["original_path"])
            if not original_path.exists():
                continue
            mat_code = file_info["material_code"]
            new_filename = PLATFORM_NAMING_RULE.format(
                product_code=code,
                scene=scene,
                date=date_str,
                material_code=mat_code
            )
            new_filename = f"{new_filename}{original_path.suffix}"
            new_path = self.output_dir / new_filename
            counter = 1
            while new_path.exists():
                new_filename = PLATFORM_NAMING_RULE.format(
                    product_code=code,
                    scene=scene,
                    date=date_str,
                    material_code=f"{mat_code}_{counter}"
                )
                new_filename = f"{new_filename}{original_path.suffix}"
                new_path = self.output_dir / new_filename
                counter += 1
            shutil.copy2(original_path, new_path)
            file_info["platform_name"] = new_filename
            file_info["platform_path"] = str(new_path)
            renamed.append(str(new_path))
        return renamed

    def _generate_submission_checklist(self, files: List[str], file_mapping: List[Dict] = None) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "提交清单"
        title = f"{self.project.product_name} - 上架材料提交清单"
        ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        current_batch = self.materials.get_current_batch()
        batch_info = ""
        if current_batch:
            batch_info = f"批次：{current_batch.batch_name} ({current_batch.batch_id})"
            ws.cell(row=2, column=1, value=batch_info).font = Font(italic=True, color="666666")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

        info_rows = [
            ["产品编码", self.project.product_code],
            ["交易场景", self.project.trading_scene],
            ["联系人", self.project.contact_person],
            ["联系电话", self.project.contact_phone],
            ["生成日期", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["项目编号", self.project.project_id]
        ]
        start_row = 4 if batch_info else 3
        for row_idx, (key, value) in enumerate(info_rows, start_row):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        start_row = start_row + len(info_rows) + 2
        headers = ["序号", "材料编码", "材料名称", "原始文件名", "平台文件名", "来源", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        summary = self.materials.get_summary()
        provided_items = []

        if file_mapping:
            for fm in file_mapping:
                mat_item = self.materials.get_item(fm["material_code"])
                if mat_item:
                    provided_items.append(mat_item)
        else:
            provided_items = self.materials.get_provided()

        for idx, item in enumerate(provided_items, 1):
            row = start_row + idx
            is_generated = item.generated

            original_name = ""
            platform_name = ""
            if file_mapping:
                for fm in file_mapping:
                    if fm["material_code"] == item.code:
                        original_name = Path(fm["original_path"]).name
                        platform_name = fm.get("platform_name", "")
                        break
            else:
                for f in files:
                    if item.code in Path(f).name:
                        platform_name = Path(f).name
                        break

            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=item.code)
            ws.cell(row=row, column=3, value=item.name)
            ws.cell(row=row, column=4, value=original_name)
            ws.cell(row=row, column=5, value=platform_name)
            ws.cell(row=row, column=6, value="自动生成" if is_generated else "用户上传")
            status_cell = ws.cell(row=row, column=7, value="已准备")
            status_cell.font = Font(color="00B050")
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border

        summary_start = start_row + len(provided_items) + 2
        ws.cell(row=summary_start, column=1, value="材料汇总").font = Font(bold=True, size=12)

        summary_data = [
            ["材料总数", summary["total"]],
            ["必填材料", summary["required"]],
            ["已提供", summary["provided"]],
            ["自动生成", summary["generated"]],
            ["缺失必填", summary["missing_required"]],
            ["缺失可选", summary["missing_optional"]],
            ["完整性", "完整" if summary["complete"] else "不完整"],
            ["批次数量", summary.get("batch_count", 0)]
        ]
        for row_idx, (key, value) in enumerate(summary_data, summary_start + 1):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            cell = ws.cell(row=row_idx, column=2, value=str(value))
            if key == "完整性" and not summary["complete"]:
                cell.font = Font(color="FF0000")
            elif key == "完整性" and summary["complete"]:
                cell.font = Font(color="#00B050")

        if self.validation_result:
            val_start = summary_start + len(summary_data) + 2
            ws.cell(row=val_start, column=1, value="校验结果").font = Font(bold=True, size=12)

            val_data = []
            if self.validation_result.errors:
                val_data.append(["错误", f"{len(self.validation_result.errors)}项"])
            if self.validation_result.warnings:
                val_data.append(["警告", f"{len(self.validation_result.warnings)}项"])
            if self.validation_result.sensitive_hits:
                val_data.append(["敏感词", f"{len(self.validation_result.sensitive_hits)}处"])

            for row_idx, (key, value) in enumerate(val_data, val_start + 1):
                ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True, color="FF0000" if key == "错误" else "FFA500")
                ws.cell(row=row_idx, column=2, value=value)

        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 22

        filename = f"{self._get_filename_prefix()}_TJQD_提交清单.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        return str(filepath)

    def _create_zip_package(self, files: List[str]) -> str:
        zip_filename = f"{self._get_filename_prefix()}_上架材料包.zip"
        zip_path = self.output_dir / zip_filename
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filepath in files:
                path_obj = Path(filepath)
                if path_obj.exists():
                    zf.write(filepath, arcname=path_obj.name)
        return str(zip_path)

    def _get_filename_prefix(self) -> str:
        date_str = datetime.now().strftime("%Y%m%d")
        scene = self._safe_filename(self.project.trading_scene[:4] if self.project.trading_scene else "场景")
        code = self._safe_filename(self.project.product_code if self.project.product_code else self.project.project_id)
        return f"{code}_{scene}_{date_str}"

    @staticmethod
    def _safe_filename(filename: str) -> str:
        invalid_chars = '<>:"/\\|?*'
        for ch in invalid_chars:
            filename = filename.replace(ch, '_')
        return filename.strip() or "unnamed"
